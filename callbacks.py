from dash import Input, Output, State, dash_table, dcc, html, clientside_callback
import dash
import pandas as pd
import numpy as np
import os
import duckdb
import plotly.express as px
import plotly.graph_objects as go
from utils.helpers import (
    eliminar_acentos, import_dict, comunas_df, puertos_coords,
    cargar_diccionarios, cargar_diccionarios_categoria_hs, cargar_descripcion_estructura,
    leer_txt_sin_encabezado, obtener_importadores_coincidentes,
    obtener_metadata_parquet, query_aggregated, query_raw, query_parquet,
    enriquecer_desde_diccionarios, listar_archivos_parquet, _create_conn, _attach_years,
    buscar_codigos_columna, get_global_conn, reset_global_conn
)
from io import StringIO
from dash.exceptions import PreventUpdate
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ── Constantes SQL compartidas ──
_product_expr = (
    "CONCAT(COALESCE(\"DNOMBRE\", ''), ' ', COALESCE(\"DMARCA\", ''), ' ', "
    "COALESCE(\"DVARIEDAD\", ''), ' ', COALESCE(\"DOTRO1\", ''), ' ', "
    "COALESCE(\"DOTRO2\", ''), ' ', COALESCE(\"ATR_5\", ''), ' ', COALESCE(\"ATR_6\", ''))"
)
_cif_expr = "\"CIF_ITEM\""
_cant_expr = "\"CANT_MERC\""

# ── Helpers de UI ──
def _card(title, children, full_width=False):
    style = {'gridColumn': '1 / -1'} if full_width else None
    return html.Div([
        html.Div(title, className='viz-card-header'),
        children
    ], className='viz-card', style=style)

_PLOTLY_THEME = dict(
    template=None,
    paper_bgcolor='rgba(0,0,0,0)',
    plot_bgcolor='rgba(0,0,0,0)',
    font=dict(size=11, color='rgba(128,128,128,0.85)'),
    xaxis=dict(gridcolor='rgba(128,128,128,0.15)', zerolinecolor='rgba(128,128,128,0.2)',
               title=dict(font=dict(color='rgba(128,128,128,0.7)'))),
    yaxis=dict(gridcolor='rgba(128,128,128,0.15)', zerolinecolor='rgba(128,128,128,0.2)',
               title=dict(font=dict(color='rgba(128,128,128,0.7)'))),
    hoverlabel=dict(font=dict(color='rgba(128,128,128,0.85)')),
    margin=dict(l=40, r=20, t=30, b=40),
)

def _apply_plotly_theme(fig):
    if fig and not isinstance(fig, str):
        try:
            fig.update_layout(**_PLOTLY_THEME)
            fig.update_layout(
                hoverlabel=dict(namelength=-1),
                yaxis=dict(tickformat=','),
            )
        except Exception:
            pass
    return fig

def _loading_graph(figure, height='400px', graph_id=None):
    _apply_plotly_theme(figure)
    graph_kwargs = {'figure': figure, 'style': {'height': height},
                    'config': {'displayModeBar': True, 'responsive': True,
                               'modeBarButtonsToRemove': ['lasso2d', 'select2d']}}
    if graph_id:
        graph_kwargs['id'] = graph_id
    return dcc.Loading(
        dcc.Graph(**graph_kwargs),
        type='circle', color='#00cec9',
        style={'height': height})

def _data_table(df, page_size=20):
    if df is None or df.empty:
        return html.P("Sin datos", style={'color': 'var(--text-muted)'})
    return dash_table.DataTable(
        data=df.to_dict('records'),
        columns=[{'name': c, 'id': c} for c in df.columns],
        style_header={'backgroundColor': 'var(--bg-hover)', 'color': 'var(--text-primary)', 'fontWeight': 'bold',
                      'borderBottom': '1px solid var(--border)'},
        style_cell={'backgroundColor': 'var(--dt-cell-bg)', 'color': 'var(--text-secondary)', 'padding': '8px',
                    'fontSize': '12px', 'textAlign': 'left'},
        style_data_conditional=[{'if': {'row_index': 'odd'}, 'backgroundColor': 'var(--bg-card)'}],
        page_size=page_size)

def _empty_fig(msg='Sin datos disponibles'):
    fig = go.Figure()
    fig.add_annotation(text=msg, xref='paper', yref='paper', x=0.5, y=0.5, showarrow=False,
                       font=dict(size=16, color='rgba(128,128,128,0.7)'))
    _apply_plotly_theme(fig)
    fig.update_layout(xaxis=dict(visible=False), yaxis=dict(visible=False))
    return fig

def _error_fig(e):
    fig = go.Figure()
    fig.add_annotation(text=f'Error: {e}', xref='paper', yref='paper', x=0.5, y=0.5, showarrow=False,
                       font=dict(size=14, color='rgba(255,107,107,0.8)'))
    _apply_plotly_theme(fig)
    fig.update_layout(xaxis=dict(visible=False), yaxis=dict(visible=False))
    return fig

# ── Filtros ──
def _componer_fecha(dia, mes, anyo):
    if dia and mes and anyo:
        return f"{anyo}-{mes}-{dia}"
    return None

def _build_filter_sql(primary_aranc, primary_importador,
                      start_day, start_month, start_year,
                      end_day, end_month, end_year,
                      search_producto, search_importador,
                      search_pa_orig, search_pa_adq, search_comuna):
    filters = []
    if primary_aranc:
        terms = [t.strip() for t in primary_aranc.split(',')]
        term_conds = [f'"ARANC_NAC" ILIKE \'{t}%\'' for t in terms if t]
        if term_conds:
            filters.append('(' + ' OR '.join(term_conds) + ')')
    if primary_importador:
        terms = [t.strip() for t in primary_importador.split(',')]
        term_conds = [f'"NUM_UNICO_IMPORTADOR" ILIKE \'%{t}%\'' for t in terms if t]
        if term_conds:
            filters.append('(' + ' OR '.join(term_conds) + ')')
    start_date = _componer_fecha(start_day, start_month, start_year)
    end_date = _componer_fecha(end_day, end_month, end_year)
    if start_date and end_date:
        filters.append(
            f"\"DD\" IS NOT NULL AND TRY_STRPTIME(\"DD\", '%d%m%Y') BETWEEN '{start_date}'::DATE AND '{end_date}'::DATE"
        )
    if search_producto:
        terms = [t.strip() for t in search_producto.split(',')]
        product_cols = ['"DNOMBRE"', '"DMARCA"', '"DVARIEDAD"', '"DOTRO1"', '"DOTRO2"', '"ATR_5"', '"ATR_6"']
        for term in terms:
            col_conditions = [f"{col} ILIKE '%{term}%'" for col in product_cols]
            filters.append('(' + ' OR '.join(col_conditions) + ')')
    if search_importador:
        terms = [t.strip() for t in search_importador.split(',')]
        term_conditions = [f'"NUM_UNICO_IMPORTADOR" ILIKE \'%{t}%\'' for t in terms if t]
        if term_conditions:
            filters.append('(' + ' OR '.join(term_conditions) + ')')
    if search_pa_orig:
        nuevos_terms = []
        for t in [x.strip() for x in search_pa_orig.split(',')]:
            codigos = buscar_codigos_columna('PA_ORIG', t)
            if codigos:
                nuevos_terms.extend(codigos)
            else:
                nuevos_terms.append(t)
        
        term_conditions = [f'"PA_ORIG" ILIKE \'%{t}%\'' for t in nuevos_terms if t]

        if term_conditions:
            filters.append('(' + ' OR '.join(term_conditions) + ')')
    if search_pa_adq:
        nuevos_terms = []
        for t in [x.strip() for x in search_pa_adq.split(',')]:
            codigos = buscar_codigos_columna('PA_ADQ', t)
            if codigos:
                nuevos_terms.extend(codigos)
            else:
                nuevos_terms.append(t)
        
        term_conditions = [f'"PA_ADQ" ILIKE \'%{t}%\'' for t in nuevos_terms if t]

        if term_conditions:
            filters.append('(' + ' OR '.join(term_conditions) + ')')

    if search_comuna:
        nuevos_terms = []
        for t in [x.strip() for x in search_comuna.split(',')]:
            codigos = buscar_codigos_columna('CODCOMUN', t)
            if codigos:
                nuevos_terms.extend(codigos)
            else:
                nuevos_terms.append(t)
        
        term_conditions = [f'"CODCOMUN" ILIKE \'%{t}%\'' for t in nuevos_terms if t]
        if term_conditions:
            filters.append('(' + ' OR '.join(term_conditions) + ')')
    return filters

def _apply_section_hs_filter(df_data, section_value, hsdesc_value):
    df_dict = cargar_diccionarios_categoria_hs()
    df_dict.columns = df_dict.columns.str.strip()
    df_dict['Chapter'] = df_dict['Chapter'].astype(str).str[:2]
    if 'ARANC_NAC' in df_data.columns:
        df_data['Chapter'] = df_data['ARANC_NAC'].astype(str).str[:2]
        df_data = df_data.merge(
            df_dict[['Chapter', 'HS Description', 'Section']].drop_duplicates(),
            on='Chapter', how='left'
        )
    if section_value:
        df_data = df_data[df_data['Section'] == section_value] if not isinstance(section_value, list) else df_data[df_data['Section'].isin(section_value)]
    if hsdesc_value:
        df_data = df_data[df_data['HS Description'] == hsdesc_value] if not isinstance(hsdesc_value, list) else df_data[df_data['HS Description'].isin(hsdesc_value)]
    return df_data

def _apply_drill_filter(filters, drill_data):
    """Applies drill-down filter to the filters list. Returns modified filters."""
    if drill_data and isinstance(drill_data, dict) and drill_data.get('active'):
        col = drill_data.get('column')
        val = drill_data.get('value')
        if col and val and col not in ('Section', 'HS Description', 'ANO', 'MES'):
            val_escaped = val.replace("'", "''")
            if col == 'PRODUCTO':
                filters.append(f'({_product_expr}) ILIKE \'%{val_escaped}%\'')
            else:
                filters.append(f'CAST("{col}" AS VARCHAR) ILIKE \'%{val_escaped}%\'')
    return filters

# ── Funciones de generación de visualizaciones ──

def _gen_monthly_charts(conn, años, where_str):
    try:
        df = query_parquet(
            "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%Y-%m') AS MES, "
            f"SUM({_cif_expr}) AS CIF_ITEM, "
            f"SUM({_cant_expr}) AS CANT_MERC",
            where_clause=where_str, group_by='1', order_by='1', conn=conn)
        if not df.empty:
            df['MES'] = pd.to_datetime(df['MES'] + '-01')
            fig_cif = px.area(df, x='MES', y='CIF_ITEM', title='CIF_ITEM Mensual vs Tiempo', template=None)
            fig_kilos = px.area(df, x='MES', y='CANT_MERC', title='CANT_MERC Mensual vs Tiempo', template=None)
            df['CIF_ITEM/KILOS'] = df['CIF_ITEM'] / df['CANT_MERC'].replace(0, np.nan)
            fig_kilos_cif = px.line(df, x='MES', y='CIF_ITEM/KILOS', title='CIF_ITEM/KILOS Mensual vs Tiempo', template=None)
            return fig_cif, fig_kilos, fig_kilos_cif, df
    except Exception as e:
        return _error_fig(e), _error_fig(e), _error_fig(e), pd.DataFrame()
    return _empty_fig(), _empty_fig(), _empty_fig(), pd.DataFrame()

def _gen_country_adq(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['MES', 'PA_ADQ'], filters,
                              extra_selects={'MES': "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%Y-%m')"}, conn=conn)
        if not df.empty:
            df['MES'] = pd.to_datetime(df['MES'] + '-01')
            df = enriquecer_desde_diccionarios(df, ['PA_ADQ'])
            return px.line(df, x='MES', y='CIF_ITEM', color='PA_ADQ',
                          title='CIF_ITEM mensual por país de adquisición', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_country_orig(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['MES', 'PA_ORIG'], filters,
                              extra_selects={'MES': "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%Y-%m')"}, conn=conn)
        if not df.empty:
            df['MES'] = pd.to_datetime(df['MES'] + '-01')
            df = enriquecer_desde_diccionarios(df, ['PA_ORIG'])
            return px.line(df, x='MES', y='CIF_ITEM', color='PA_ORIG',
                          title='CIF_ITEM mensual por país de origen', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_pct_bar(conn, años, column_dropdown, filters):
    try:
        df = query_aggregated('CIF_ITEM', [column_dropdown], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, [column_dropdown])
            if column_dropdown == 'NUM_UNICO_IMPORTADOR':
                df[column_dropdown] = df[column_dropdown].astype(str).map(import_dict).fillna(df[column_dropdown].astype(str))
            total = df['CIF_ITEM'].sum()
            df['%'] = (df['CIF_ITEM'] / total * 100).round(2)
            df = df.sort_values('CIF_ITEM', ascending=False)
            fig = px.bar(df, x=column_dropdown, y='%',
                         title=f'Porcentaje de CIF_ITEM por {column_dropdown}', template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_section_pie(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['ARANC_NAC'], filters, conn=conn)
        if not df.empty:
            df = _apply_section_hs_filter(df, None, None)
            if 'Section' in df.columns:
                g = df.groupby('Section')['CIF_ITEM'].sum().reset_index()
                return px.pie(g, values='CIF_ITEM', names='Section',
                              title='Porcentaje del Tipo de Producto (Section)', template=None)
    except Exception:
        import traceback; traceback.print_exc()
        return None

def _gen_top20_analysis(conn, años, where_str):
    """Merged query for top20 freq + val (same GROUP BY)."""
    try:
        df = query_parquet(
            f"{_product_expr} AS PRODUCTO, COUNT(*) AS Conteo, SUM({_cif_expr}) AS Total_CIF",
            where_clause=where_str, group_by="PRODUCTO", conn=conn)
        if df is None or df.empty:
            empty = pd.DataFrame(columns=['PRODUCTO', 'Conteo'])
            return _empty_fig(), empty, _empty_fig(), pd.DataFrame(columns=['PRODUCTO', 'Total_CIF'])

        freq_df = df.nlargest(20, 'Conteo').reset_index(drop=True)
        val_df = df.nlargest(20, 'Total_CIF').reset_index(drop=True)

        fig_freq = _empty_fig()
        if not freq_df.empty:
            fig_freq = px.bar(freq_df, x='Conteo', y='PRODUCTO', orientation='h',
                             title='Top 20 Productos (frecuencia)', template=None)
            fig_freq.update_layout(yaxis={'categoryorder': 'total ascending', 'automargin': True})

        fig_val = _empty_fig()
        if not val_df.empty:
            fig_val = px.bar(val_df, x='Total_CIF', y='PRODUCTO', orientation='h',
                            title='Top 20 Productos (valor CIF)', template=None)
            fig_val.update_layout(yaxis={'categoryorder': 'total ascending', 'automargin': True})
        return fig_freq, freq_df, fig_val, val_df
    except Exception as e:
        return (_error_fig(e), pd.DataFrame(columns=['PRODUCTO', 'Conteo']),
                _error_fig(e), pd.DataFrame(columns=['PRODUCTO', 'Total_CIF']))

def _gen_top20_ind(conn, años, filters):
    cols = ['DNOMBRE', 'DMARCA', 'DVARIEDAD', 'DOTRO1', 'DOTRO2', 'ATR_5', 'ATR_6',
            'TPO_DOCTO', 'ARANC_NAC', 'NUM_UNICO_IMPORTADOR', 'CIF_ITEM', 'CANT_MERC',
            'DESOBS1', 'DD', 'CODCOMUN', 'ADU', 'PTO_DESEM', 'PTO_EMB', 'VIA_TRAN']
    try:
        df = query_raw(columns=cols, filters=filters, limit=2000, conn=conn)
        if df is not None and not df.empty:
            df = enriquecer_desde_diccionarios(df, ['CODCOMUN', 'ADU', 'PTO_DESEM', 'PTO_EMB', 'VIA_TRAN'])
            df['NUM_UNICO_IMPORTADOR'] = df['NUM_UNICO_IMPORTADOR'].astype(str).map(import_dict).fillna(df['NUM_UNICO_IMPORTADOR'].astype(str))
            for c in ['CIF_ITEM', 'CANT_MERC']:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors='coerce')
            product_parts = [df[c].fillna('') for c in ['DNOMBRE', 'DMARCA', 'DVARIEDAD', 'DOTRO1', 'DOTRO2', 'ATR_5', 'ATR_6']]
            df['PRODUCTO'] = product_parts[0]
            for p in product_parts[1:]:
                df['PRODUCTO'] = df['PRODUCTO'] + ' ' + p
            return df.nlargest(20, 'CIF_ITEM')[
                ['PRODUCTO', 'TPO_DOCTO', 'ARANC_NAC', 'NUM_UNICO_IMPORTADOR',
                 'CIF_ITEM', 'CANT_MERC', 'DESOBS1', 'DD', 'CODCOMUN', 'ADU',
                 'PTO_DESEM', 'PTO_EMB', 'VIA_TRAN']]
    except Exception:
        import traceback; traceback.print_exc()
        return pd.DataFrame()

def _gen_avg_price_orig(conn, años, where_str):
    try:
        w = where_str + f' AND {_cif_expr} IS NOT NULL' if where_str else f'{_cif_expr} IS NOT NULL'
        df = query_parquet(
            f"CAST(\"PA_ORIG\" AS VARCHAR) AS PA_ORIG, SUM({_cif_expr}) AS CIF_ITEM, SUM({_cant_expr}) AS CANT_MERC",
            where_clause=w, group_by='"PA_ORIG"', conn=conn)
        if not df.empty:
            pp = df.copy()
            pp['Precio Promedio (CIF/Kg)'] = (pp['CIF_ITEM'] / pp['CANT_MERC'].replace(0, np.nan)).round(2)
            pp = enriquecer_desde_diccionarios(pp[['PA_ORIG', 'Precio Promedio (CIF/Kg)']], ['PA_ORIG'])
            return pp
    except Exception:
        import traceback; traceback.print_exc()
        return pd.DataFrame(columns=['PA_ORIG', 'Precio Promedio (CIF/Kg)'])

def _gen_avg_price_adq(conn, años, where_str):
    try:
        w = where_str + f' AND {_cif_expr} IS NOT NULL' if where_str else f'{_cif_expr} IS NOT NULL'
        df = query_parquet(
            f"CAST(\"PA_ADQ\" AS VARCHAR) AS PA_ADQ, SUM({_cif_expr}) AS CIF_ITEM, SUM({_cant_expr}) AS CANT_MERC",
            where_clause=w, group_by='"PA_ADQ"', conn=conn)
        if not df.empty:
            pp = df.copy()
            pp['Precio Promedio (CIF/Kg)'] = (pp['CIF_ITEM'] / pp['CANT_MERC'].replace(0, np.nan)).round(2)
            pp = enriquecer_desde_diccionarios(pp[['PA_ADQ', 'Precio Promedio (CIF/Kg)']], ['PA_ADQ'])
            return pp
    except Exception:
        import traceback; traceback.print_exc()
        return pd.DataFrame(columns=['PA_ADQ', 'Precio Promedio (CIF/Kg)'])

def _gen_heat_origen(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['PA_ORIG', 'MES'], filters,
                              extra_selects={'MES': "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%Y-%m')"}, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['PA_ORIG'])
            df['MES'] = pd.to_datetime(df['MES'] + '-01')
            pivot = df.pivot_table(index='PA_ORIG', columns='MES', values='CIF_ITEM', aggfunc='sum', fill_value=0)
            fig = px.imshow(pivot, title='Heatmap País Origen vs Tiempo', template=None)
            fig.update_yaxes(automargin=True, tickfont=dict(size=9))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_heat_adq(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['PA_ADQ', 'MES'], filters,
                              extra_selects={'MES': "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%Y-%m')"}, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['PA_ADQ'])
            df['MES'] = pd.to_datetime(df['MES'] + '-01')
            pivot = df.pivot_table(index='PA_ADQ', columns='MES', values='CIF_ITEM', aggfunc='sum', fill_value=0)
            fig = px.imshow(pivot, title='Heatmap País Adquisición vs Tiempo', template=None)
            fig.update_yaxes(automargin=True, tickfont=dict(size=9))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_mapa_comunas(conn, años, where_str):
    try:
        w = where_str + f' AND {_cif_expr} IS NOT NULL' if where_str else f'{_cif_expr} IS NOT NULL'
        df = query_parquet(
            f"CAST(\"CODCOMUN\" AS VARCHAR) AS CODCOMUN, SUM({_cif_expr}) AS \"Total CIF_ITEM\", SUM({_cant_expr}) AS \"Total Cantidad\"",
            where_clause=w, group_by='"CODCOMUN"', conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['CODCOMUN'])
            df['CODCOMUN'] = df['CODCOMUN'].astype(str).str.strip().str.lower().apply(eliminar_acentos)
            loc = comunas_df.copy()
            loc['nombre'] = loc['nombre'].str.strip().str.lower().apply(eliminar_acentos)
            merged = df.merge(loc[['nombre', 'latitud', 'longitud']], left_on='CODCOMUN', right_on='nombre', how='left').dropna(subset=['latitud', 'longitud'])
            fig = px.scatter_mapbox(merged, lat='latitud', lon='longitud', hover_name='CODCOMUN',
                                    color='Total CIF_ITEM', size='Total CIF_ITEM',
                                    title='Mapa de Comunas', zoom=4, height=800, template=None)
            fig.update_layout(mapbox_style="carto-positron")
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_port_analysis(conn, años, where_str):
    """Merged query for port charts (mapa_puertos + sankey + port_matrix)."""
    try:
        df = query_parquet(
            f'CAST("PTO_EMB" AS VARCHAR) AS PTO_EMB, CAST("PTO_DESEM" AS VARCHAR) AS PTO_DESEM, {_cif_expr} AS CIF_ITEM',
            where_clause=where_str, order_by='CIF_ITEM DESC', limit=5000, conn=conn)
        if df.empty:
            return _empty_fig(), _empty_fig(), _empty_fig()

        df = enriquecer_desde_diccionarios(df, ['PTO_EMB', 'PTO_DESEM'])

        # mapa_puertos
        loc = puertos_coords.copy()
        loc['Puerto'] = loc['Puerto'].str.strip().str.lower().apply(eliminar_acentos)
        emb = df.groupby('PTO_EMB')['CIF_ITEM'].sum().reset_index().rename(columns={'PTO_EMB': 'Puerto'})
        des = df.groupby('PTO_DESEM')['CIF_ITEM'].sum().reset_index().rename(columns={'PTO_DESEM': 'Puerto'})
        all_p = pd.concat([emb, des]).groupby('Puerto')['CIF_ITEM'].sum().reset_index()
        all_p['key'] = all_p['Puerto'].str.strip().str.lower().apply(eliminar_acentos)
        merged = all_p.merge(loc, left_on='key', right_on='Puerto', how='inner')
        if not merged.empty:
            merged['Latitud'] = pd.to_numeric(merged['Latitud'], errors='coerce')
            fig_mapa = px.scatter_mapbox(merged, lat='Latitud', lon='Longitud',
                                         hover_name='Puerto_x', size='CIF_ITEM', color='CIF_ITEM',
                                         title='Mapa de Puertos (volumen CIF)', zoom=5, height=800,
                                         center={'lat': -33.45, 'lon': -70.65}, template=None)
            fig_mapa.update_layout(mapbox_style="carto-positron")
        else:
            fig_mapa = _empty_fig()

        # sankey
        if 'PTO_EMB' in df.columns and 'PTO_DESEM' in df.columns:
            agg = df.groupby(['PTO_EMB', 'PTO_DESEM'])['CIF_ITEM'].sum().reset_index()
            agg = agg.sort_values('CIF_ITEM', ascending=False).head(50)
            all_ports = pd.unique(agg[['PTO_EMB', 'PTO_DESEM']].values.ravel())
            port_idx = {p: i for i, p in enumerate(all_ports)}
            fig_sankey = go.Figure(data=[go.Sankey(
                node=dict(label=list(all_ports), pad=15, thickness=20, color='#00b894'),
                link=dict(
                    source=[port_idx[s] for s in agg['PTO_EMB']],
                    target=[port_idx[t] for t in agg['PTO_DESEM']],
                    value=agg['CIF_ITEM']
                )
            )])
            fig_sankey.update_layout(title='Flujo de Puertos (Embarque → Desembarque)', template=None)
        else:
            fig_sankey = _empty_fig()

        # port_matrix
        pivot = df.pivot_table(index='PTO_EMB', columns='PTO_DESEM', values='CIF_ITEM', aggfunc='sum', fill_value=0)
        pivot = pivot.loc[pivot.sum(axis=1).nlargest(10).index, pivot.sum(axis=0).nlargest(10).index]
        fig_matrix = px.imshow(pivot, title='Matriz Puerto Embarque → Desembarque (Top 10)',
                              template=None, aspect='auto',
                              labels=dict(x='Puerto Desembarque', y='Puerto Embarque', color='CIF_ITEM'))

        return fig_mapa, fig_sankey, fig_matrix
    except Exception as e:
        return _error_fig(e), _error_fig(e), _error_fig(e)

def _gen_importadores(conn, años, filters):
    try:
        df = query_raw(columns=['NUM_UNICO_IMPORTADOR'], filters=filters, limit=50000, conn=conn)
        if df is not None and not df.empty:
            df['NUM_UNICO_IMPORTADOR_ORIGINAL'] = df['NUM_UNICO_IMPORTADOR']
            df['NUM_UNICO_IMPORTADOR'] = df['NUM_UNICO_IMPORTADOR'].apply(lambda x: import_dict.get(str(x).strip(), x))
            return obtener_importadores_coincidentes(df)
    except Exception:
        import traceback; traceback.print_exc()
        return pd.DataFrame(columns=['RUT_ORIGINAL', 'NOMBRE_REEMPLAZADO'])

def _gen_indicadores(df_mensual):
    try:
        total_cif = df_mensual['CIF_ITEM'].sum() if not df_mensual.empty else 0
        total_kilos = df_mensual['CANT_MERC'].sum() if not df_mensual.empty else 0
        precio_kg = total_cif / total_kilos if total_kilos else 0
        return total_cif, total_kilos, precio_kg
    except:
        return 0, 0, 0

def _gen_yoy(conn, años, where_str):
    try:
        df = query_parquet(
            "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%m') AS MES_NUM, "
            "STRFTIME(TRY_STRPTIME(\"DD\", '%d%m%Y'), '%Y') AS ANO, "
            f"SUM({_cif_expr}) AS CIF_ITEM",
            where_clause=where_str, group_by='1, 2', order_by='2, 1', conn=conn)
        if not df.empty:
            df['MES_NUM'] = df['MES_NUM'].astype(str).str.zfill(2)
            return px.line(df, x='MES_NUM', y='CIF_ITEM', color='ANO',
                          title='Comparación Interanual (YTD por mes)', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_transporte(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['VIA_TRAN'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['VIA_TRAN'])
            df = df.sort_values('CIF_ITEM', ascending=False)
            return px.bar(df, x='VIA_TRAN', y='CIF_ITEM',
                         title='CIF_ITEM por Medio de Transporte', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_aduana(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['ADU'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['ADU'])
            df = df.sort_values('CIF_ITEM', ascending=False).head(15)
            fig = px.bar(df, x='ADU', y='CIF_ITEM', title='Top 15 Aduanas por CIF_ITEM', template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_price_hist(conn, años, where_str):
    try:
        w = (where_str + ' AND ' if where_str else '') + f'{_cif_expr} > 0 AND {_cant_expr} > 0'
        df = query_parquet(
            f"{_cif_expr} / NULLIF({_cant_expr}, 0) AS PRECIO_UNIT",
            where_clause=w, order_by='PRECIO_UNIT DESC', limit=5000, conn=conn)
        if not df.empty and 'PRECIO_UNIT' in df.columns:
            df['PRECIO_UNIT'] = pd.to_numeric(df['PRECIO_UNIT'], errors='coerce')
            df = df.dropna(subset=['PRECIO_UNIT'])
            df = df[np.isfinite(df['PRECIO_UNIT'])]
            df = df[df['PRECIO_UNIT'] < df['PRECIO_UNIT'].quantile(0.99)]
            if df.empty:
                return _empty_fig()
            return px.histogram(df, x='PRECIO_UNIT', nbins=50,
                               title='Distribución de Precios Unitarios (CIF/Kg)', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_operacion(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['TPO_DOCTO'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['TPO_DOCTO'])
            df = df.sort_values('CIF_ITEM', ascending=False)
            return px.pie(df, values='CIF_ITEM', names='TPO_DOCTO',
                         title='Distribución por Tipo de Operación', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_bultos(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['TPO_BUL1'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['TPO_BUL1'])
            df = df.sort_values('CIF_ITEM', ascending=False).head(15)
            fig = px.bar(df, x='TPO_BUL1', y='CIF_ITEM', title='Top 15 Tipos de Bulto por CIF', template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_box_precios(conn, años, where_str):
    try:
        w = (where_str + ' AND ' if where_str else '') + f'{_cif_expr} > 0'
        df = query_parquet(
            f"CAST(\"PA_ORIG\" AS VARCHAR) AS PA_ORIG, {_cif_expr} / NULLIF({_cant_expr}, 0) AS PRECIO_UNIT",
            where_clause=w, order_by='PRECIO_UNIT DESC', limit=10000, conn=conn)
        if not df.empty and 'PA_ORIG' in df.columns:
            df['PRECIO_UNIT'] = pd.to_numeric(df['PRECIO_UNIT'], errors='coerce').dropna()
            df = df[df['PRECIO_UNIT'] < df['PRECIO_UNIT'].quantile(0.99)]
            df = enriquecer_desde_diccionarios(df, ['PA_ORIG'])
            top = df.groupby('PA_ORIG')['PRECIO_UNIT'].count().nlargest(15).index
            df = df[df['PA_ORIG'].isin(top)]
            fig = px.box(df, x='PA_ORIG', y='PRECIO_UNIT',
                        title='Distribución de Precios Unitarios por País de Origen', template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_treemap(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['ARANC_NAC'], filters,
                              extra_selects={'DNOMBRE': 'COALESCE(MIN(CAST("DNOMBRE" AS VARCHAR)), MIN(CAST("ARANC_NAC" AS VARCHAR)))'}, conn=conn)
        if not df.empty:
            df = _apply_section_hs_filter(df, None, None)
            agg = df.groupby(['Section', 'HS Description', 'ARANC_NAC', 'DNOMBRE'], as_index=False)['CIF_ITEM'].sum().dropna(subset=['Section'])
            agg['DNOMBRE'] = agg['DNOMBRE'].fillna(agg['ARANC_NAC'])
            fig = px.treemap(agg, path=['Section', 'HS Description', 'ARANC_NAC'], values='CIF_ITEM',
                             title='Jerarquía de Productos (Section → HS → Código Arancelario)', template=None)
            fig.update_traces(hovertemplate='<b>%{label}</b><br>Producto: %{customdata[0]}<br>CIF: %{value:,.0f}<extra></extra>',
                              customdata=agg[['DNOMBRE']].values)
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

# ── Nuevos gráficos (diccionarios Comext) ──

def _gen_monedas(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['MONEDA'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['MONEDA'])
            df = df.sort_values('CIF_ITEM', ascending=False).head(15)
            fig = px.bar(df, x='MONEDA', y='CIF_ITEM', title='CIF por Moneda',
                         template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_formas_pago(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['FORM_PAGO'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['FORM_PAGO'])
            df = df.sort_values('CIF_ITEM', ascending=False)
            return px.bar(df, x='FORM_PAGO', y='CIF_ITEM',
                         title='CIF por Forma de Pago', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_clausula(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['CL_COMPRA'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['CL_COMPRA'])
            df = df.sort_values('CIF_ITEM', ascending=False)
            return px.pie(df, values='CIF_ITEM', names='CL_COMPRA',
                         title='Distribución por Cláusula de Compra (Incoterms)',
                         template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_origen_divisas(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['CODORDIV'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['CODORDIV'])
            df = df.sort_values('CIF_ITEM', ascending=False)
            return px.bar(df, x='CODORDIV', y='CIF_ITEM',
                         title='CIF por Origen de Divisas', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_regimen_importacion(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['REG_IMP'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['REG_IMP'])
            df = df.sort_values('CIF_ITEM', ascending=False)
            return px.bar(df, x='REG_IMP', y='CIF_ITEM',
                         title='CIF por Régimen de Importación', template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig('Columna REG_IMP sin datos en estos años')

def _gen_unidad_medida(conn, años, filters):
    try:
        df = query_aggregated('CANT_MERC', ['MEDIDA'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['MEDIDA'])
            df = df.sort_values('CANT_MERC', ascending=False).head(15)
            fig = px.bar(df, x='MEDIDA', y='CANT_MERC',
                         title='Top 15 Unidades de Medida por Cantidad', template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_bancos(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['BCO_COM'], filters, conn=conn)
        if not df.empty:
            df = enriquecer_desde_diccionarios(df, ['BCO_COM'])
            df = df.sort_values('CIF_ITEM', ascending=False).head(15)
            fig = px.bar(df, x='BCO_COM', y='CIF_ITEM',
                         title='Top 15 Bancos Comerciales por CIF', template=None)
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_cost_breakdown(conn, años, where_str):
    try:
        df = query_parquet(
            f"SUM({_cif_expr}) AS CIF_ITEM, SUM(CAST(REPLACE(CAST(\"FOB\" AS VARCHAR), ',', '.') AS DOUBLE)) AS FOB, SUM(CAST(REPLACE(CAST(\"FLETE\" AS VARCHAR), ',', '.') AS DOUBLE)) AS FLETE, SUM(CAST(REPLACE(CAST(\"SEGURO\" AS VARCHAR), ',', '.') AS DOUBLE)) AS SEGURO",
            where_clause=where_str, conn=conn)
        if not df.empty and len(df) == 1:
            row = df.iloc[0]
            costs = pd.DataFrame({
                'Componente': ['FOB (mercancía)', 'FLETE (flete)', 'SEGURO (seguro)', 'Diferencia CIF'],
                'Valor': [row.get('FOB', 0) or 0, row.get('FLETE', 0) or 0, row.get('SEGURO', 0) or 0,
                          max(0, (row.get('CIF_ITEM', 0) or 0) - (row.get('FOB', 0) or 0) - (row.get('FLETE', 0) or 0) - (row.get('SEGURO', 0) or 0))]
            })
            fig = px.bar(costs, x='Componente', y='Valor', title='Desglose de Costos (FOB + Flete + Seguro vs CIF)',
                         template=None, text_auto='.0f')
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            fig.update_traces(textposition='outside')
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_importer_concentration(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['NUM_UNICO_IMPORTADOR'], filters, conn=conn)
        if not df.empty:
            total = df['CIF_ITEM'].sum()
            df = df.sort_values('CIF_ITEM', ascending=False)
            df['% Acumulado'] = (df['CIF_ITEM'].cumsum() / total * 100).round(1)
            df['% Individual'] = (df['CIF_ITEM'] / total * 100).round(1)
            top10 = df.head(10).copy()
            top10['Importador'] = top10['NUM_UNICO_IMPORTADOR'].astype(str).map(import_dict).fillna(top10['NUM_UNICO_IMPORTADOR'].astype(str))
            fig = px.bar(top10, x='Importador', y='% Individual',
                         title='Top 10 Importadores (% del CIF Total)',
                         template=None, text='% Individual')
            fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
            fig.update_traces(texttemplate='%{text}%', textposition='outside')
            return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_dispatch_days(conn, años, where_str):
    try:
        df = query_parquet(
            "CAST(\"FECACEP\" AS VARCHAR) AS FECACEP, CAST(\"FEC_ALMAC\" AS VARCHAR) AS FEC_ALMAC",
            where_clause=where_str, limit=5000, conn=conn)
        if not df.empty:
            for c in ['FECACEP', 'FEC_ALMAC']:
                df[c] = pd.to_datetime(df[c], format='%d%m%Y', errors='coerce')
            df = df.dropna(subset=['FECACEP', 'FEC_ALMAC'])
            df['Días Despacho'] = (df['FEC_ALMAC'] - df['FECACEP']).dt.days
            df = df[(df['Días Despacho'] >= 0) & (df['Días Despacho'] < 365)]
            if not df.empty:
                return px.histogram(df, x='Días Despacho', nbins=50,
                                   title='Distribución de Días de Despacho (FECACEP → FEC_ALMAC)',
                                   template=None)
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

def _gen_tariff_analysis(conn, años, filters):
    try:
        df = query_aggregated('CIF_ITEM', ['ARANC_NAC'], filters, conn=conn)
        if not df.empty:
            df = _apply_section_hs_filter(df, None, None)
            if 'Section' in df.columns:
                w = ' AND '.join(filters) if filters else None
                tariff_df = query_parquet(
                    f"CAST(\"ARANC_NAC\" AS VARCHAR) AS ARANC_NAC, CAST(REPLACE(CAST(\"ADVAL_ALA\" AS VARCHAR), ',', '.') AS DOUBLE) AS ADVAL_ALA, {_cif_expr} AS CIF_ITEM",
                    where_clause=w, limit=10000, conn=conn)
                if not tariff_df.empty:
                    tariff_df = _apply_section_hs_filter(tariff_df, None, None)
                    avg = tariff_df.groupby('Section').apply(
                        lambda x: (x['ADVAL_ALA'] * x['CIF_ITEM']).sum() / x['CIF_ITEM'].sum() if x['CIF_ITEM'].sum() > 0 else 0
                    ).reset_index(name='Arancel Promedio (%)')
                    avg = avg[avg['Arancel Promedio (%)'] > 0].sort_values('Arancel Promedio (%)', ascending=False)
                    if not avg.empty:
                        fig = px.bar(avg, x='Section', y='Arancel Promedio (%)',
                                    title='Arancel Efectivo Promedio por Sección',
                                    template=None, text_auto='.2f')
                        fig.update_layout(xaxis_tickangle=-45, xaxis=dict(automargin=True))
                        fig.update_traces(texttemplate='%{text}%', textposition='outside')
                        return fig
    except Exception as e:
        return _error_fig(e)
    return _empty_fig()

# ── Funciones merge ──

def _gen_country_analysis(conn, años, filters):
    try:
        fig_orig = _gen_country_orig(conn, años, filters)
        fig_adq = _gen_country_adq(conn, años, filters)
        return fig_orig, fig_adq
    except Exception as e:
        return _error_fig(e), _error_fig(e)

def _gen_heat_analysis(conn, años, filters):
    try:
        fig_orig = _gen_heat_origen(conn, años, filters)
        fig_adq = _gen_heat_adq(conn, años, filters)
        return fig_orig, fig_adq
    except Exception as e:
        return _error_fig(e), _error_fig(e)

def _gen_avg_price_analysis(conn, años, where_str):
    try:
        pp_orig = _gen_avg_price_orig(conn, años, where_str)
        pp_adq = _gen_avg_price_adq(conn, años, where_str)
        return pp_orig, pp_adq
    except:
        import traceback; traceback.print_exc()
        return (pd.DataFrame(columns=['PA_ORIG', 'Precio Promedio (CIF/Kg)']),
                pd.DataFrame(columns=['PA_ADQ', 'Precio Promedio (CIF/Kg)']))


# ── Cache de tabs y mapping gen → tab ──
_TAB_CACHE_LOCK = threading.Lock()
_TAB_CACHE = {}
_TAB_GENS = {
    'Resumen': ['monthly', 'yoy', 'price_hist', 'pct_bar', 'importer_conc'],
    'Paises': ['country_analysis', 'heat_analysis', 'box_precios'],
    'Productos': ['section', 'top20_analysis', 'treemap'],
    'Transporte y Rutas': ['transporte', 'aduana', 'operacion', 'bultos', 'port_analysis'],
    'Geografía': ['mapa_comunas', 'port_analysis'],
    'Tablas': ['top20_analysis', 'top20_ind', 'avg_price_analysis', 'importadores'],
    'Financiero': ['monedas', 'formas_pago', 'clausula', 'origen_divisas', 'cost_breakdown'],
    'Clasificación': ['regimen', 'unidad', 'bancos', 'tariff', 'dispatch_days'],
}

# Mapping gen_key → (func, extra_args_fn) donde extra_args_fn(filters, where_str, column_dropdown) → args
_GEN_CALLS = {
    'monthly':        (_gen_monthly_charts,  lambda f, w, c: [w]),
    'country_analysis':(_gen_country_analysis,lambda f, w, c: [f]),
    'heat_analysis':  (_gen_heat_analysis,   lambda f, w, c: [f]),
    'pct_bar':        (_gen_pct_bar,         lambda f, w, c: [c, f]),
    'section':        (_gen_section_pie,     lambda f, w, c: [f]),
    'top20_analysis': (_gen_top20_analysis,  lambda f, w, c: [w]),
    'top20_ind':      (_gen_top20_ind,       lambda f, w, c: [f]),
    'avg_price_analysis':(_gen_avg_price_analysis, lambda f, w, c: [w]),
    'mapa_comunas':   (_gen_mapa_comunas,    lambda f, w, c: [w]),
    'port_analysis':  (_gen_port_analysis,   lambda f, w, c: [w]),
    'importadores':   (_gen_importadores,    lambda f, w, c: [f]),
    'yoy':            (_gen_yoy,             lambda f, w, c: [w]),
    'transporte':     (_gen_transporte,      lambda f, w, c: [f]),
    'aduana':         (_gen_aduana,          lambda f, w, c: [f]),
    'price_hist':     (_gen_price_hist,      lambda f, w, c: [w]),
    'operacion':      (_gen_operacion,       lambda f, w, c: [f]),
    'bultos':         (_gen_bultos,          lambda f, w, c: [f]),
    'box_precios':    (_gen_box_precios,     lambda f, w, c: [w]),
    'treemap':        (_gen_treemap,         lambda f, w, c: [f]),
    'monedas':        (_gen_monedas,         lambda f, w, c: [f]),
    'formas_pago':    (_gen_formas_pago,     lambda f, w, c: [f]),
    'clausula':       (_gen_clausula,        lambda f, w, c: [f]),
    'origen_divisas': (_gen_origen_divisas,  lambda f, w, c: [f]),
    'regimen':        (_gen_regimen_importacion, lambda f, w, c: [f]),
    'unidad':         (_gen_unidad_medida,   lambda f, w, c: [f]),
    'bancos':         (_gen_bancos,          lambda f, w, c: [f]),
    'cost_breakdown': (_gen_cost_breakdown,  lambda f, w, c: [w]),
    'importer_conc':  (_gen_importer_concentration, lambda f, w, c: [f]),
    'dispatch_days':  (_gen_dispatch_days,   lambda f, w, c: [w]),
    'tariff':         (_gen_tariff_analysis, lambda f, w, c: [f]),
}

_tab_style = {
    'backgroundColor': 'var(--bg-secondary)', 'color': 'var(--text-secondary)',
    'border': '1px solid var(--border)', 'borderBottom': 'none',
    'padding': '10px 18px', 'fontWeight': 'bold', 'fontSize': '13px'
}
_tab_selected_style = {
    'backgroundColor': 'var(--bg-primary)', 'color': 'var(--accent)',
    'border': '1px solid var(--border)', 'borderBottom': 'none',
    'borderTop': '2px solid var(--accent)', 'padding': '10px 18px',
    'fontWeight': 'bold', 'fontSize': '13px'
}
_tabs_colors = {'border': 'var(--border)', 'primary': 'var(--accent)', 'background': 'var(--bg-primary)'}
_tab_labels = ['Resumen', 'Paises', 'Productos', 'Transporte y Rutas',
               'Geografía', 'Tablas', 'Financiero', 'Clasificación']

_grid = {'display': 'grid', 'gridTemplateColumns': '1fr 1fr', 'gap': '16px'}


def _run_gens(años, where_str, filters, column_dropdown, gen_keys):
    if not gen_keys:
        return {}

    results = {}
    max_workers = min(len(gen_keys), 4)

    def _run_one(key):
        func, args_fn = _GEN_CALLS[key]
        extra = args_fn(filters, where_str, column_dropdown)
        conn = _create_conn(años)
        try:
            return key, func(conn, años, *extra)
        finally:
            try:
                conn.close()
            except Exception:
                pass

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_key = {pool.submit(_run_one, key): key for key in gen_keys}
        for future in as_completed(future_to_key):
            key = future_to_key[future]
            try:
                _, result = future.result()
                results[key] = result
            except Exception:
                results[key] = _error_fig(f'Error en {key}')

    return results


def _build_tab(tab_name, results):
    """Build tab content from gen results dict."""
    if tab_name == 'Resumen':
        fig_m, fig_mk, fig_mkc, df_m = results['monthly']
        tc, tk, pk = _gen_indicadores(df_m)
        card_indicadores = html.Div([
            html.Div([
                html.Div("Total CIF", className='stat-label'),
                html.Div(f"{tc:,.0f}", style={'color': 'var(--accent)'}, className='stat-value')
            ], className='stat-item'),
            html.Div([
                html.Div("Total Kg", className='stat-label'),
                html.Div(f"{tk:,.0f}", style={'color': 'var(--success)'}, className='stat-value')
            ], className='stat-item'),
            html.Div([
                html.Div("Precio por Kg", className='stat-label'),
                html.Div(f"${pk:,.2f}", style={'color': 'var(--danger)'}, className='stat-value')
            ], className='stat-item'),
        ], className='stat-card')
        return html.Div([
            html.Div(card_indicadores, style={'gridColumn': '1 / -1'}),
            _card("Evolucion Mensual CIF", _loading_graph(fig_m, graph_id='g-m-cif')),
            _card("Evolucion Mensual Cantidad", _loading_graph(fig_mk, graph_id='g-m-kilos')),
            _card("CIF / Kilos Mensual", _loading_graph(fig_mkc, graph_id='g-m-kilos-cif')),
            _card("Comparacion Interanual por Mes", _loading_graph(results['yoy'], graph_id='g-yoy')),
            _card("Distribucion de Precios Unitarios", _loading_graph(results['price_hist'], graph_id='g-precios')),
            _card("Porcentaje por Variable", _loading_graph(results['pct_bar'], graph_id='g-pct')),
            _card("Concentracion de Importadores", _loading_graph(results['importer_conc'], graph_id='g-importer-conc')),
        ], style=_grid)

    if tab_name == 'Paises':
        fig_orig, fig_adq = results['country_analysis']
        heat_orig, heat_adq = results['heat_analysis']
        return html.Div([
            _card("CIF Mensual por Pais de Origen", _loading_graph(fig_orig, graph_id='g-orig')),
            _card("CIF Mensual por Pais de Adquisicion", _loading_graph(fig_adq, graph_id='g-adq')),
            _card("Heatmap Pais Origen vs Tiempo", _loading_graph(heat_orig, graph_id='g-heat-orig'), full_width=True),
            _card("Heatmap Pais Adquisicion vs Tiempo", _loading_graph(heat_adq, graph_id='g-heat-adq'), full_width=True),
            _card("Distribucion de Precios por Pais", _loading_graph(results['box_precios'], graph_id='g-box-precios'), full_width=True),
        ], style=_grid)

    if tab_name == 'Productos':
        fig_s = results['section']
        fig_tf, df_freq, fig_tv, df_val = results['top20_analysis']
        return html.Div([
            _card("Distribucion por Seccion (Section)",
                  _loading_graph(fig_s if fig_s else _empty_fig(), graph_id='g-section')),
            _card("Top 20 Productos por Frecuencia", _loading_graph(fig_tf, graph_id='g-top20-freq'), full_width=True),
            _card("Top 20 Productos por Valor CIF", _loading_graph(fig_tv, graph_id='g-top20-val'), full_width=True),
            _card("Jerarquia de Productos (Treemap)", _loading_graph(results['treemap'], graph_id='g-treemap'), full_width=True),
        ], style=_grid)

    if tab_name == 'Transporte y Rutas':
        fig_mapa, fig_sankey, fig_matrix = results['port_analysis']
        return html.Div([
            _card("Medio de Transporte", _loading_graph(results['transporte'], graph_id='g-transporte')),
            _card("Aduanas", _loading_graph(results['aduana'], graph_id='g-aduana')),
            _card("Tipo de Operacion", _loading_graph(results['operacion'], graph_id='g-operacion')),
            _card("Top 15 Tipos de Bulto", _loading_graph(results['bultos'], graph_id='g-bultos')),
            _card("Diagrama Sankey: Embarque -> Desembarque",
                  dcc.Loading(dcc.Graph(id='g-sankey', figure=fig_sankey, style={'height': '700px'}, config={'displayModeBar': True}),
                              type='circle', color='#00cec9'), full_width=True),
            _card("Matriz Puertos (Embarque x Desembarque)",
                  _loading_graph(fig_matrix, graph_id='g-port-matrix'), full_width=True),
        ], style=_grid)

    if tab_name == 'Geografía':
        fig_mapa, _, _ = results['port_analysis']
        return html.Div([
            _card("Mapa de Comunas",
                  dcc.Loading(dcc.Graph(id='g-comunas', figure=results['mapa_comunas'], style={'height': '600px'}, config={'displayModeBar': True}),
                              type='circle', color='#00cec9'), full_width=True),
            _card("Mapa de Puertos",
                  dcc.Loading(dcc.Graph(id='g-puertos', figure=fig_mapa, style={'height': '600px'}, config={'displayModeBar': True}),
                              type='circle', color='#00cec9'), full_width=True),
        ], style=_grid)

    if tab_name == 'Tablas':
        _, top20, _, top20_trans = results['top20_analysis']
        pp_orig, pp_adq = results['avg_price_analysis']
        return html.Div([
            _card("Importadores Coincidentes",
                  _data_table(results['importadores'], page_size=200) if not results['importadores'].empty
                  else html.P("No hay importadores coincidentes", style={'color': 'var(--danger)'})),
            _card("Top 20 Productos por Frecuencia (Completo)", _data_table(top20)),
            _card("Top 20 Productos por Valor CIF (Completo)", _data_table(top20_trans)),
            _card("Top 20 Transacciones Individuales", _data_table(results['top20_ind'])),
            _card("Precio Promedio por Pais de Origen", _data_table(pp_orig)),
            _card("Precio Promedio por Pais de Adquisicion", _data_table(pp_adq)),
        ], style={'display': 'grid', 'gridTemplateColumns': '1fr', 'gap': '16px'})

    if tab_name == 'Financiero':
        return html.Div([
            _card("CIF por Moneda", _loading_graph(results['monedas'], graph_id='g-monedas')),
            _card("CIF por Forma de Pago", _loading_graph(results['formas_pago'], graph_id='g-formas-pago')),
            _card("Cláusula de Compra (Incoterms)", _loading_graph(results['clausula'], graph_id='g-clausula')),
            _card("CIF por Origen de Divisas", _loading_graph(results['origen_divisas'], graph_id='g-origen-divisas')),
            _card("Desglose de Costos (FOB / Flete / Seguro)", _loading_graph(results['cost_breakdown'], graph_id='g-cost-breakdown'), full_width=True),
        ], style=_grid)

    if tab_name == 'Clasificación':
        return html.Div([
            _card("CIF por Régimen de Importación", _loading_graph(results['regimen'], graph_id='g-regimen')),
            _card("Top 15 Unidades de Medida", _loading_graph(results['unidad'], graph_id='g-unidad')),
            _card("Top 15 Bancos Comerciales", _loading_graph(results['bancos'], graph_id='g-bancos')),
            _card("Arancel Efectivo Promedio por Sección", _loading_graph(results['tariff'], graph_id='g-tariff'), full_width=True),
            _card("Días de Despacho (Aceptación → Almacenaje)", _loading_graph(results['dispatch_days'], graph_id='g-dispatch')),
        ], style=_grid)

    return html.P("Tab desconocido")


def _generar_visualizaciones(años, primary_aranc, primary_importador,
                              start_day, start_month, start_year,
                              end_day, end_month, end_year,
                              search_producto, search_importador, search_pa_orig,
                              search_pa_adq, search_comuna, column_dropdown,
                              section_value, hsdesc_value, drill_data=None):
    """Genera TODOS los tabs (para exportación HTML). Reusa _TAB_CACHE si está disponible."""
    import hashlib, json
    key_parts = [str(x) for x in [años, primary_aranc, primary_importador,
                start_day, start_month, start_year,
                end_day, end_month, end_year,
                search_producto, search_importador,
                search_pa_orig, search_pa_adq, search_comuna, column_dropdown,
                section_value, hsdesc_value, drill_data]]
    cur_key = hashlib.md5('|'.join(key_parts).encode()).hexdigest()

    filters = _build_filter_sql(primary_aranc, primary_importador,
                                start_day, start_month, start_year,
                                end_day, end_month, end_year,
                                search_producto, search_importador,
                                search_pa_orig, search_pa_adq, search_comuna)
    filters = _apply_drill_filter(filters, drill_data)
    where_str = ' AND '.join(filters) if filters else None

    # Reuse cached tabs when possible
    tabs = []
    gen_needed = []
    for label in _tab_labels:
        with _TAB_CACHE_LOCK:
            cached = _TAB_CACHE.get(label)
        if cached and cached[0] == cur_key:
            tabs.append((label, cached[1]))
        else:
            gen_needed.append(label)
            tabs.append((label, None))

    if gen_needed:
        needed_keys = list({k for label in gen_needed for k in _TAB_GENS[label]})
        results = _run_gens(años, where_str, filters, column_dropdown, needed_keys)
        for label in gen_needed:
            tab_results = {k: results[k] for k in _TAB_GENS[label]}
            content = _build_tab(label, tab_results)
            with _TAB_CACHE_LOCK:
                _TAB_CACHE[label] = (cur_key, content)
            # Replace None in tabs
            idx = _tab_labels.index(label)
            tabs[idx] = (label, content)

    tab_children = [
        dcc.Tab(label=label, style=_tab_style, selected_style=_tab_selected_style, children=content)
        for label, content in tabs
    ]
    return dcc.Tabs(tab_children, style={'marginTop': '10px'}, colors=_tabs_colors)


def _precache_next_tab(active_tab, cur_key, años, where_str, filters, column_dropdown):
    """Pre-genera el siguiente tab en background."""
    idx = _tab_labels.index(active_tab)
    next_idx = (idx + 1) % len(_tab_labels)
    next_tab = _tab_labels[next_idx]
    with _TAB_CACHE_LOCK:
        if next_tab in _TAB_CACHE and _TAB_CACHE[next_tab][0] == cur_key:
            return
    gen_keys = _TAB_GENS.get(next_tab, [])
    if not gen_keys:
        return
    try:
        results = _run_gens(años, where_str, filters, column_dropdown, gen_keys)
        content = _build_tab(next_tab, results)
        with _TAB_CACHE_LOCK:
            _TAB_CACHE[next_tab] = (cur_key, content)
    except Exception:
        import traceback; traceback.print_exc()


# ── Callbacks ──
def register_callbacks(app):
    @app.callback(
        Output('stored-data', 'data'),
        Output('selected-years', 'data'),
        Output('carga-status', 'children'),
        Output('start-day', 'value'),
        Output('start-month', 'value'),
        Output('start-year', 'value'),
        Output('end-day', 'value'),
        Output('end-month', 'value'),
        Output('end-year', 'value'),
        Output('search-producto', 'disabled'),
        Output('search-importador', 'disabled'),
        Output('search-pa-orig', 'disabled'),
        Output('search-pa-adq', 'disabled'),
        Output('search-comuna', 'disabled'),
        Output('column-dropdown', 'disabled'),
        Input('cargar-button', 'n_clicks'),
        State('year-checklist', 'value'),
        State('primary-aranc', 'value'),
        State('primary-importador', 'value'),
        prevent_initial_call=True
    )
    def cargar_desde_parquet(n_clicks, años_seleccionados, primary_aranc, primary_importador):
        if not n_clicks or not años_seleccionados:
            raise PreventUpdate
        try:
            from utils.helpers import _db_exists
            dbs = [a for a in años_seleccionados if _db_exists(a)]
            if not dbs:
                return (None, None, "No hay base DuckDB para los años seleccionados. Ejecuta scripts/construir_base_duckdb.py primero.",
                        None, None, None, None, None, None, True, True, True, True, True, True)
            metadata = obtener_metadata_parquet(años_seleccionados)
            min_d = metadata.get('min_date')
            max_d = metadata.get('max_date')
            return (
                {'source': 'parquet', 'sections': metadata.get('sections', []),
                 'hs_descriptions': metadata.get('hs_descriptions', []),
                 'primary_aranc': primary_aranc, 'primary_importador': primary_importador},
                años_seleccionados,
                f"Listo ({len(dbs)} años cargados)",
                min_d[8:10] if min_d else None,
                min_d[5:7] if min_d else None,
                min_d[0:4] if min_d else None,
                max_d[8:10] if max_d else None,
                max_d[5:7] if max_d else None,
                max_d[0:4] if max_d else None,
                False, False, False, False, False, False
            )
        except Exception as e:
            import traceback; traceback.print_exc()
            return (None, None, f"Error: {e}",
                    None, None, None, None, None, None, True, True, True, True, True, True)

    # Ingesta incremental
    @app.callback(
        Output('ingesta-status', 'children'),
        Input('btn-ingestar', 'n_clicks'),
        prevent_initial_call=True
    )
    def ejecutar_ingesta(n_clicks):
        if not n_clicks:
            raise PreventUpdate
        try:
            import subprocess, sys as _sys
            script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scripts', 'ingestar.py')
            result = subprocess.run([_sys.executable, script], capture_output=True, text=True, timeout=600)
            output = result.stdout.strip().split('\n')[-2:]
            if result.returncode == 0:
                return f'OK: {output[-1]}' if output else 'Ingesta completada'
            else:
                return f'Error: {result.stderr[:200]}'
        except subprocess.TimeoutExpired:
            return 'Error: tiempo excedido (10 min)'
        except Exception as e:
            return f'Error: {e}'

    # Drill-down: click en grafico -> filtrar ----
    _DRILL_MAP = {
        'g-section': 'Section', 'g-treemap': 'Section',
        'g-top20-freq': 'PRODUCTO', 'g-top20-val': 'PRODUCTO',
        'g-orig': 'PA_ORIG', 'g-heat-orig': 'PA_ORIG', 'g-box-precios': 'PA_ORIG',
        'g-adq': 'PA_ADQ', 'g-heat-adq': 'PA_ADQ',
        'g-transporte': 'VIA_TRAN', 'g-aduana': 'ADU',
        'g-operacion': 'TPO_DOCTO', 'g-bultos': 'TPO_BUL1',
        'g-comunas': 'CODCOMUN',
        'g-yoy': 'ANO', 'g-m-cif': 'MES', 'g-m-kilos': 'MES',
        'g-monedas': 'MONEDA', 'g-formas-pago': 'FORM_PAGO',
        'g-clausula': 'CL_COMPRA', 'g-origen-divisas': 'CODORDIV',
        'g-regimen': 'REG_IMP', 'g-unidad': 'MEDIDA',
        'g-bancos': 'BCO_COM', 'g-tariff': 'Section',
    }

    @app.callback(
        Output('drill-store', 'data'),
        Output('drill-indicator', 'children'),
        Output('drill-indicator', 'style'),
        [Input('g-section', 'clickData'),
         Input('g-top20-freq', 'clickData'),
         Input('g-top20-val', 'clickData'),
         Input('g-orig', 'clickData'),
         Input('g-adq', 'clickData'),
         Input('g-transporte', 'clickData'),
         Input('g-aduana', 'clickData'),
         Input('g-operacion', 'clickData'),
         Input('g-bultos', 'clickData'),
         Input('g-comunas', 'clickData')],
        prevent_initial_call=True
    )
    def handle_drill(*args):
        ctx = dash.callback_context
        if not ctx.triggered:
            raise PreventUpdate
        trig_id = ctx.triggered[0]['prop_id'].split('.')[0]
        click_data = ctx.triggered[0]['value']
        if not click_data or 'points' not in click_data or not click_data['points']:
            raise PreventUpdate
        pt = click_data['points'][0]
        label = pt.get('label') or pt.get('x') or pt.get('y') or pt.get('text')
        if not label or label == 'Sin datos':
            raise PreventUpdate
        col = _DRILL_MAP.get(trig_id)
        if not col:
            raise PreventUpdate
        drill = {'active': True, 'column': col, 'value': str(label)}
        indicator = html.Div([
            html.Span(f"Filtro activo: {col} = {label}  ", className='drill-label'),
            html.Button("X Limpiar", id='btn-clear-drill', n_clicks=0, className='drill-clear-btn')
        ], className='drill-bar')
        return drill, indicator, {'display': 'block'}

    @app.callback(
        Output('drill-store', 'data', allow_duplicate=True),
        Output('drill-indicator', 'children', allow_duplicate=True),
        Output('drill-indicator', 'style', allow_duplicate=True),
        Input('btn-clear-drill', 'n_clicks'),
        prevent_initial_call=True
    )
    def clear_drill(n_clicks):
        return {'active': False, 'column': None, 'value': None}, '', {'display': 'none'}

    @app.callback(
        Output('sidebar', 'style'),
        Output('main-content', 'style'),
        Input('sidebar-toggle', 'n_clicks'),
        State('sidebar-state', 'data'),
        prevent_initial_call=True
    )
    def toggle_sidebar(n_clicks, state):
        if not n_clicks:
            raise PreventUpdate
        if state == 'expanded':
            new_state = 'collapsed'
            sidebar_style = {'width': '48px', 'display': 'inline-block', 'verticalAlign': 'top',
                             'marginRight': '10px', 'overflow': 'hidden'}
            content_style = {'width': 'calc(100% - 78px)', 'display': 'inline-block',
                             'verticalAlign': 'top'}
        else:
            new_state = 'expanded'
            sidebar_style = {'width': '300px', 'display': 'inline-block', 'verticalAlign': 'top',
                             'marginRight': '25px'}
            content_style = {'width': 'calc(100% - 345px)', 'display': 'inline-block',
                             'verticalAlign': 'top'}
        return sidebar_style, content_style

    @app.callback(
        Output('sidebar-state', 'data'),
        Input('sidebar-toggle', 'n_clicks'),
        State('sidebar-state', 'data'),
        prevent_initial_call=True
    )
    def save_sidebar_state(n_clicks, state):
        if not n_clicks:
            raise PreventUpdate
        return 'collapsed' if state == 'expanded' else 'expanded'

    @app.callback(
        Output('output-visualizations', 'children'),
        Output('viz-cache', 'data'),
        Output('viz-cache-key', 'data'),
        Input('stored-data', 'data'),
        Input('selected-years', 'data'),
        Input('start-day', 'value'),
        Input('start-month', 'value'),
        Input('start-year', 'value'),
        Input('end-day', 'value'),
        Input('end-month', 'value'),
        Input('end-year', 'value'),
        Input('search-producto', 'value'),
        Input('search-importador', 'value'),
        Input('search-pa-orig', 'value'),
        Input('search-pa-adq', 'value'),
        Input('search-comuna', 'value'),
        Input('column-dropdown', 'value'),
        Input('section-dropdown', 'value'),
        Input('hsdesc-dropdown', 'value'),
        Input('drill-store', 'data'),
        Input('main-tabs', 'value'),
        State('viz-cache-key', 'data'),
        prevent_initial_call=True
    )
    def update_visualizations(stored_data, selected_years,
                               start_day, start_month, start_year,
                               end_day, end_month, end_year,
                               search_producto, search_importador, search_pa_orig,
                               search_pa_adq, search_comuna, column_dropdown,
                               section_value, hsdesc_value, drill_data, active_tab, prev_key):
        if not selected_years:
            return (html.Div([html.H3("Selecciona años y presiona 'Cargar Datos'",
                                      style={'color': 'var(--text-muted)', 'textAlign': 'center'})]),
                    None, None)
        primary_aranc = primary_importador = None
        if stored_data and isinstance(stored_data, dict):
            primary_aranc = stored_data.get('primary_aranc')
            primary_importador = stored_data.get('primary_importador')
        import hashlib, json
        key_parts = [str(x) for x in [selected_years, primary_aranc, primary_importador,
                    start_day, start_month, start_year, end_day, end_month, end_year,
                    search_producto, search_importador,
                    search_pa_orig, search_pa_adq, search_comuna, column_dropdown,
                    section_value, hsdesc_value, drill_data]]
        cur_key = hashlib.md5('|'.join(key_parts).encode()).hexdigest()
        with _TAB_CACHE_LOCK:
            if cur_key == prev_key and active_tab in _TAB_CACHE and _TAB_CACHE[active_tab][0] == cur_key:
                return _TAB_CACHE[active_tab][1], {'tab': active_tab}, cur_key

        filters = _build_filter_sql(primary_aranc, primary_importador,
                                    start_day, start_month, start_year,
                                    end_day, end_month, end_year,
                                    search_producto, search_importador,
                                    search_pa_orig, search_pa_adq, search_comuna)

        filters = _apply_drill_filter(filters, drill_data)
        where_str = ' AND '.join(filters) if filters else None

        gen_keys = _TAB_GENS.get(active_tab, [])
        results = _run_gens(selected_years, where_str, filters, column_dropdown, gen_keys)
        content = _build_tab(active_tab, results)

        with _TAB_CACHE_LOCK:
            _TAB_CACHE[active_tab] = (cur_key, content)

        # Pre-cache siguiente tab en background
        t = threading.Thread(target=_precache_next_tab,
                             args=(active_tab, cur_key, selected_years,
                                   where_str, filters, column_dropdown),
                             daemon=True)
        t.start()

        return content, {'tab': active_tab}, cur_key

    @app.callback(
        Output('download-csv', 'data'),
        Input('btn-export-csv', 'n_clicks'),
        State('selected-years', 'data'),
        State('stored-data', 'data'),
        State('start-day', 'value'),
        State('start-month', 'value'),
        State('start-year', 'value'),
        State('end-day', 'value'),
        State('end-month', 'value'),
        State('end-year', 'value'),
        State('search-producto', 'value'),
        State('search-importador', 'value'),
        State('search-pa-orig', 'value'),
        State('search-pa-adq', 'value'),
        State('search-comuna', 'value'),
        prevent_initial_call=True
    )
    def export_csv(n_clicks, selected_years, stored_data,
                   start_day, start_month, start_year,
                   end_day, end_month, end_year,
                   search_producto, search_importador,
                   search_pa_orig, search_pa_adq, search_comuna):
        if not n_clicks or not selected_years:
            raise PreventUpdate
        primary_aranc = primary_importador = None
        if stored_data and isinstance(stored_data, dict):
            primary_aranc = stored_data.get('primary_aranc')
            primary_importador = stored_data.get('primary_importador')
        filters = _build_filter_sql(primary_aranc, primary_importador,
                                    start_day, start_month, start_year,
                                    end_day, end_month, end_year,
                                    search_producto, search_importador,
                                    search_pa_orig, search_pa_adq, search_comuna)
        conn = get_global_conn(selected_years)
        cols = ['DD', 'ARANC_NAC', 'NUM_UNICO_IMPORTADOR', 'CIF_ITEM', 'CANT_MERC',
                'PA_ORIG', 'PA_ADQ', 'CODCOMUN', 'VIA_TRAN', 'DNOMBRE', 'DMARCA']
        df = query_raw(columns=cols, filters=filters, limit=50000, conn=conn)
        if df is not None and not df.empty:
            df = enriquecer_desde_diccionarios(df, ['PA_ORIG', 'PA_ADQ', 'CODCOMUN', 'VIA_TRAN'])
            if 'NUM_UNICO_IMPORTADOR' in df.columns:
                df['NUM_UNICO_IMPORTADOR'] = df['NUM_UNICO_IMPORTADOR'].astype(str).map(import_dict).fillna(df['NUM_UNICO_IMPORTADOR'].astype(str))
            return dcc.send_string(df.to_csv(index=False, encoding='utf-8-sig'), "importaciones_export.csv") 
        raise PreventUpdate

    @app.callback(
        Output('download-html', 'data'),
        Input('btn-export-html', 'n_clicks'),
        State('selected-years', 'data'),
        State('stored-data', 'data'),
        State('start-day', 'value'),
        State('start-month', 'value'),
        State('start-year', 'value'),
        State('end-day', 'value'),
        State('end-month', 'value'),
        State('end-year', 'value'),
        State('search-producto', 'value'),
        State('search-importador', 'value'),
        State('search-pa-orig', 'value'),
        State('search-pa-adq', 'value'),
        State('search-comuna', 'value'),
        State('column-dropdown', 'value'),
        State('section-dropdown', 'value'),
        State('hsdesc-dropdown', 'value'),
        prevent_initial_call=True
    )
    def export_html(n_clicks, selected_years, stored_data,
                    start_day, start_month, start_year,
                    end_day, end_month, end_year,
                    search_producto, search_importador,
                    search_pa_orig, search_pa_adq, search_comuna,
                    column_dropdown, section_value, hsdesc_value):
        if not n_clicks or not selected_years:
            raise PreventUpdate
        primary_aranc = primary_importador = None
        if stored_data and isinstance(stored_data, dict):
            primary_aranc = stored_data.get('primary_aranc')
            primary_importador = stored_data.get('primary_importador')
        tabs = _generar_visualizaciones(
            selected_years, primary_aranc, primary_importador,
            start_day, start_month, start_year,
            end_day, end_month, end_year,
            search_producto, search_importador,
            search_pa_orig, search_pa_adq, search_comuna,
            column_dropdown, section_value, hsdesc_value
        )
        import plotly.io as pio
        try:
            html_parts = ['<html><head><meta charset="utf-8"><title>Dashboard Importaciones</title>',
                          '<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script></head><body>']
            excluded = {'Importadores Coincidentes', 'Top 20 Productos por Frecuencia (Completo)',
                        'Top 20 Productos por Valor CIF (Completo)', 'Top 20 Transacciones Individuales',
                        'Precio Promedio por Pais de Origen', 'Precio Promedio por Pais de Adquisicion'}
            for tab in tabs.children:
                content_div = tab.children
                if hasattr(content_div, 'children') and isinstance(content_div.children, list):
                    label = getattr(tab, 'label', '')
                    html_parts.append(f'<h2>{label}</h2>')
                    for child in content_div.children:
                        if hasattr(child, 'children') and isinstance(child.children, list):
                            title = ''
                            graph = None
                            for c in child.children:
                                if isinstance(c, str):
                                    title = c
                                if hasattr(c, 'type') and c.type == 'Loading':
                                    if hasattr(c, 'children') and hasattr(c.children, 'type') and c.children.type == 'Graph':
                                        graph = c.children.figure
                            if title and graph and title not in excluded:
                                try:
                                    html_parts.append(f'<h3>{title}</h3>')
                                    html_parts.append(pio.to_html(graph, include_plotlyjs=False, full_html=False))
                                except:
                                    pass
            html_parts.append('</body></html>')
            content = '\n'.join(html_parts)
            return dcc.send_string(content, "dashboard_graficos.html")
        except Exception as e:
            import traceback; traceback.print_exc()
            raise PreventUpdate

    @app.callback(
        Output('download-excel', 'data'),
        Input('btn-export-excel', 'n_clicks'),
        State('selected-years', 'data'),
        State('stored-data', 'data'),
        State('start-day', 'value'),
        State('start-month', 'value'),
        State('start-year', 'value'),
        State('end-day', 'value'),
        State('end-month', 'value'),
        State('end-year', 'value'),
        State('search-producto', 'value'),
        State('search-importador', 'value'),
        State('search-pa-orig', 'value'),
        State('search-pa-adq', 'value'),
        State('search-comuna', 'value'),
        prevent_initial_call=True
    )
    def export_excel(n_clicks, selected_years, stored_data,
                     start_day, start_month, start_year,
                     end_day, end_month, end_year,
                     search_producto, search_importador,
                     search_pa_orig, search_pa_adq, search_comuna):
        if not n_clicks or not selected_years:
            raise PreventUpdate
        primary_aranc = primary_importador = None
        if stored_data and isinstance(stored_data, dict):
            primary_aranc = stored_data.get('primary_aranc')
            primary_importador = stored_data.get('primary_importador')
        filters = _build_filter_sql(primary_aranc, primary_importador,
                                    start_day, start_month, start_year,
                                    end_day, end_month, end_year,
                                    search_producto, search_importador,
                                    search_pa_orig, search_pa_adq, search_comuna)
        where_str = ' AND '.join(filters) if filters else None
        conn = get_global_conn(selected_years)
        cols = ['DD', 'ARANC_NAC', 'NUM_UNICO_IMPORTADOR', 'CIF_ITEM', 'CANT_MERC',
                 'PA_ORIG', 'PA_ADQ', 'CODCOMUN', 'VIA_TRAN', 'DNOMBRE', 'DMARCA',
                 'TPO_DOCTO', 'PTO_EMB', 'PTO_DESEM', 'ADU']
        df = query_raw(columns=cols, filters=filters, limit=50000, conn=conn)
        if df is not None and not df.empty:
            df = enriquecer_desde_diccionarios(df, ['PA_ORIG', 'PA_ADQ', 'CODCOMUN', 'VIA_TRAN',
                                                     'PTO_EMB', 'PTO_DESEM', 'ADU', 'TPO_DOCTO'])
            if 'NUM_UNICO_IMPORTADOR' in df.columns:
                df['NUM_UNICO_IMPORTADOR'] = df['NUM_UNICO_IMPORTADOR'].astype(str).map(import_dict).fillna(df['NUM_UNICO_IMPORTADOR'].astype(str))
            import openpyxl
            import io
            from openpyxl.styles import Font, PatternFill, Alignment
            buf = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Importaciones"
            header_fill = PatternFill(start_color='00CEC9', end_color='00CEC9', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for c, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=c, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
            for r, row in df.iterrows():
                for c, val in enumerate(row, 1):
                    ws.cell(row=r+2, column=c, value=val)
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
            wb.save(buf)
            buf.seek(0)
            return dcc.send_bytes(buf.read(), "importaciones_export.xlsx")
        raise PreventUpdate

    @app.callback(
        Output('section-dropdown', 'options'),
        Input('stored-data', 'data'),
        Input('selected-years', 'data')
    )
    def update_section_options(stored_data, selected_years):
        if selected_years and stored_data and isinstance(stored_data, dict):
            return [{'label': str(s), 'value': str(s)} for s in stored_data.get('sections', [])]
        return []

    @app.callback(
        [Output('hsdesc-dropdown', 'options'), Output('hsdesc-dropdown', 'value')],
        [Input('stored-data', 'data'), Input('selected-years', 'data'), Input('section-dropdown', 'value')],
        [State('hsdesc-dropdown', 'value')]
    )
    def update_hsdesc_options(stored_data, selected_years, section_value, hsdesc_value):
        hs_descs = []
        if selected_years and stored_data and isinstance(stored_data, dict):
            hs_descs = stored_data.get('hs_descriptions', [])
        if section_value and hs_descs:
            try:
                df_dict = cargar_diccionarios_categoria_hs()
                df_dict.columns = df_dict.columns.str.strip()
                if isinstance(section_value, list):
                    df_dict = df_dict[df_dict['Section'].isin(section_value)]
                else:
                    df_dict = df_dict[df_dict['Section'] == section_value]
                hs_descs = sorted(df_dict['HS Description'].dropna().unique().tolist())
            except Exception:
                import traceback; traceback.print_exc()
        return [{'label': str(h), 'value': str(h)} for h in hs_descs] if hs_descs else [], None

    # Autocomplete para importador (busqueda DuckDB contra archivo TSV)
    @app.callback(
        Output('primary-importador', 'options'),
        Input('primary-importador', 'search_value'),
        prevent_initial_call=True
    )
    def autocomplete_importador(search_value):
        if not search_value or len(search_value.strip()) < 2:
            return []
        try:
            sv = search_value.strip().replace("'", "''")
            conn2 = _create_conn()
            df = conn2.execute(f"""
                SELECT RUT, RAZON_SOCIAL FROM read_csv_auto('data/import.txt', delim='\t', header=true, ignore_errors=true)
                WHERE CAST(RUT AS VARCHAR) ILIKE '%{sv}%'
                   OR CAST(RAZON_SOCIAL AS VARCHAR) ILIKE '%{sv}%'
                LIMIT 50
            """).fetchdf()
            conn2.close()
            return [{'label': f"{r['RUT']} - {str(r.get('RAZON_SOCIAL', '')).strip()}", 'value': str(r['RUT'])}
                    for _, r in df.iterrows() if r['RUT'] is not None]
        except:
            return []

    clientside_callback(
        """
        function(n_clicks) {
            if (n_clicks > 0) {
                setTimeout(function() { window.print(); }, 500);
            }
            return 0;
        }
        """,
        Output('btn-print', 'n_clicks'),
        Input('btn-print', 'n_clicks'),
        prevent_initial_call=True
    )

    # ── Loading overlay show/hide ──
    clientside_callback(
        """
        function(tab, stored) {
            var el = document.getElementById('loading-overlay');
            if (el) el.classList.remove('hidden');
            return window.dash_clientside.no_update;
        }
        """,
        Output('main-tabs', 'id'),
        Input('main-tabs', 'value'),
        Input('stored-data', 'data'),
        prevent_initial_call=True
    )

    clientside_callback(
        """
        function(children) {
            var el = document.getElementById('loading-overlay');
            if (el && children && children.props) el.classList.add('hidden');
            return window.dash_clientside.no_update;
        }
        """,
        Output('output-visualizations', 'id'),
        Input('output-visualizations', 'children'),
        prevent_initial_call=True
    )

    # ── Theme toggle ──
    clientside_callback(
        """
        function(n_clicks, current) {
            var newTheme = current === 'dark' ? 'light' : 'dark';
            document.documentElement.setAttribute('data-theme', newTheme);
            localStorage.setItem('theme', newTheme);
            // Update Plotly chart text/grid colors
            var color = newTheme === 'dark' ? '#c8d6e5' : '#636e72';
            var grid = newTheme === 'dark' ? '#2d3138' : '#d1d8e0';
            document.querySelectorAll('.js-plotly-plot').forEach(function(el) {
                if (window.Plotly) Plotly.relayout(el, {
                    'font.color': color,
                    'xaxis.gridcolor': grid,
                    'yaxis.gridcolor': grid,
                    'xaxis.zerolinecolor': grid,
                    'yaxis.zerolinecolor': grid,
                });
            });
            return newTheme;
        }
        """,
        Output('theme-store', 'data'),
        Input('btn-theme-toggle', 'n_clicks'),
        State('theme-store', 'data'),
        prevent_initial_call=True
    )

    # Update toggle button icon based on theme
    clientside_callback(
        """
        function(theme) {
            return theme === 'dark' ? '☀' : '☾';
        }
        """,
        Output('btn-theme-toggle', 'children'),
        Input('theme-store', 'data'),
    )
